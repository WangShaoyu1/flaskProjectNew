#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel模板生成工具
用于生成指令和词槽的Excel导入模板
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_instruction_template():
    """创建指令导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "指令导入模板"
    
    # 定义列标题 - 按照402文件夹实际格式
    headers = [
        "分类", "指令名称", "指令标识", "指令描述", 
        "关联词槽", "追问话术", "追问失败话术", "追问次数",
        "相似问", "执行成功话术", "执行失败话术"
    ]
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    desc_font = Font(italic=True, color="666666")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入标题行
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 添加示例数据 - 按照402文件夹实际格式
    example_data = [
        [
            "设备控制", "开灯指令", "LIGHT_ON", "打开灯光设备",
            "设备名称,房间位置", "请问您要打开哪个房间的灯？", "抱歉，我没有理解您的意思", "2",
            "开灯|打开灯|亮灯|开启灯光", "灯光已为您打开", "抱歉，灯光打开失败，请稍后重试"
        ],
        [
            "设备控制", "关灯指令", "LIGHT_OFF", "关闭灯光设备",
            "设备名称,房间位置", "请问您要关闭哪个房间的灯？", "抱歉，我没有理解您的意思", "2",
            "关灯|关闭灯|熄灯|关闭灯光", "灯光已为您关闭", "抱歉，灯光关闭失败，请稍后重试"
        ],
        [
            "信息查询", "查询天气", "WEATHER_QUERY", "查询天气信息",
            "地点,日期", "请问您要查询哪个地方的天气？", "抱歉，我没有理解您的意思", "2",
            "天气怎么样|今天天气|天气预报|查天气", "今天天气晴朗，温度适宜", "抱歉，天气查询失败，请稍后重试"
        ]
    ]
    
    for row, data in enumerate(example_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
    
    # 调整列宽
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 特殊列宽调整
    ws.column_dimensions['C'].width = 20  # instruction_desc
    ws.column_dimensions['G'].width = 20  # success_response
    ws.column_dimensions['H'].width = 20  # failure_response
    ws.column_dimensions['K'].width = 25  # similar_questions
    
    # 添加说明工作表
    ws_info = wb.create_sheet("填写说明")
    info_data = [
        ["字段名", "说明", "示例"],
        ["instruction_name", "指令名称，必填", "开灯指令"],
        ["instruction_code", "指令代码，必填，唯一", "LIGHT_ON"],
        ["instruction_desc", "指令描述", "打开灯光设备"],
        ["category", "分类", "设备控制"],
        ["is_slot_related", "是否关联词槽，TRUE/FALSE", "TRUE"],
        ["related_slot_ids", "关联词槽ID，多个用逗号分隔", "device_id,room_id"],
        ["success_response", "成功响应", "灯光已打开"],
        ["failure_response", "失败响应", "灯光打开失败"],
        ["is_enabled", "是否启用，TRUE/FALSE", "TRUE"],
        ["sort_order", "排序号", "1"],
        ["similar_questions", "相似问题，多个用|分隔", "开灯|打开灯|亮灯"]
    ]
    
    for row, data in enumerate(info_data, 1):
        for col, value in enumerate(data, 1):
            cell = ws_info.cell(row=row, column=col, value=value)
            if row == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.border = border
    
    # 调整说明工作表列宽
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 30
    ws_info.column_dimensions['C'].width = 25
    
    return wb


def create_slot_template():
    """创建词槽导入模板"""
    wb = Workbook()
    
    # 修改默认工作表
    ws = wb.active
    ws.title = "词槽"
    
    # 词槽列标题 - 按照402文件夹实际格式
    headers = [
        "词槽名称", "词槽描述", "实体ID", "实体标准名", "实体别名"
    ]
    
    # 设置样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    desc_font = Font(italic=True, color="666666")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入标题行
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 添加示例数据 - 按照402文件夹实际格式
    example_data = [
        ["休眠时间", "", 1, "15秒", "十五秒"],
        ["", "", 2, "30秒", "三十秒"],
        ["", "", 3, "1分钟", "1分==一分钟==一分"],
        ["", "", 4, "2分钟", "2分==两分钟==二分钟==两分"],
        ["", "", 5, "5分钟", "5分==五分钟==五分"],
        ["设备名称", "智能家居设备名称", 6, "灯", "灯光==电灯==照明灯"],
        ["", "", 7, "空调", "空调机==冷气==AC"],
        ["", "", 8, "电视", "电视机==TV==电视机"],
        ["房间位置", "房间或区域位置", 9, "客厅", "大厅==起居室==客厅"],
        ["", "", 10, "卧室", "卧房==睡房==房间"],
        ["", "", 11, "厨房", "厨房==烹饪间"]
    ]
    
    for row, data in enumerate(example_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15  # 词槽名称
    ws.column_dimensions['B'].width = 20  # 词槽描述
    ws.column_dimensions['C'].width = 10  # 实体ID
    ws.column_dimensions['D'].width = 15  # 实体标准名
    ws.column_dimensions['E'].width = 30  # 实体别名
    
    # 添加说明工作表
    ws_info = wb.create_sheet("填写说明")
    info_data = [
        ["工作表", "字段名", "说明", "示例"],
        ["词槽定义", "slot_name", "词槽名称，必填", "设备ID"],
        ["词槽定义", "slot_code", "词槽代码，必填，唯一", "device_id"],
        ["词槽定义", "slot_type", "词槽类型", "text"],
        ["词槽定义", "slot_desc", "词槽描述", "设备标识符"],
        ["词槽定义", "is_required", "是否必填，TRUE/FALSE", "TRUE"],
        ["词槽定义", "default_value", "默认值", ""],
        ["词槽定义", "validation_rule", "验证规则", ""],
        ["词槽定义", "is_enabled", "是否启用，TRUE/FALSE", "TRUE"],
        ["词槽定义", "sort_order", "排序号", "1"],
        ["词槽值", "slot_code", "词槽代码，必须与词槽定义中的匹配", "device_id"],
        ["词槽值", "standard_value", "标准值，必填", "light_001"],
        ["词槽值", "synonyms", "同义词，多个用|分隔", "灯1|主灯|客厅灯"],
        ["词槽值", "value_desc", "值描述", "客厅主灯"],
        ["词槽值", "is_enabled", "是否启用，TRUE/FALSE", "TRUE"],
        ["词槽值", "sort_order", "排序号", "1"]
    ]
    
    for row, data in enumerate(info_data, 1):
        for col, value in enumerate(data, 1):
            cell = ws_info.cell(row=row, column=col, value=value)
            if row == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.border = border
    
    # 调整说明工作表列宽
    ws_info.column_dimensions['A'].width = 12
    ws_info.column_dimensions['B'].width = 20
    ws_info.column_dimensions['C'].width = 30
    ws_info.column_dimensions['D'].width = 25
    
    return wb


def main():
    """主函数"""
    # 创建模板目录
    template_dir = "templates"
    if not os.path.exists(template_dir):
        os.makedirs(template_dir)
    
    print("开始生成Excel模板...")
    
    # 生成指令模板
    try:
        instruction_wb = create_instruction_template()
        instruction_path = os.path.join(template_dir, "指令导入模板.xlsx")
        instruction_wb.save(instruction_path)
        print(f"✓ 指令导入模板已生成: {instruction_path}")
    except Exception as e:
        print(f"✗ 生成指令模板失败: {e}")
    
    # 生成词槽模板
    try:
        slot_wb = create_slot_template()
        slot_path = os.path.join(template_dir, "词槽导入模板.xlsx")
        slot_wb.save(slot_path)
        print(f"✓ 词槽导入模板已生成: {slot_path}")
    except Exception as e:
        print(f"✗ 生成词槽模板失败: {e}")
    
    print("模板生成完成！")
    print(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main() 