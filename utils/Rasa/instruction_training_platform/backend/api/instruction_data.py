"""
智能对话训练平台 - 指令数据管理API
提供指令数据的CRUD操作、相似问管理和批量操作功能
"""

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List, Optional, Dict, Any
import json
import pandas as pd
from io import BytesIO
import tempfile
import os

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(__file__)))

from models.database_models import (
    get_new_db, 
    InstructionData as DBInstructionData,
    SimilarQuestion as DBSimilarQuestion,
    InstructionLibraryMaster as DBInstructionLibraryMaster,
    SlotDefinition as DBSlotDefinition
)
from models.schemas import (
    InstructionData, InstructionDataCreate, 
    SimilarQuestion, SimilarQuestionCreate,
    ApiResponse
)
from response_utils import StandardResponse, success_response, error_response, ErrorCodes, Messages

router = APIRouter(prefix="/api/v2/instruction", tags=["指令数据管理"])


@router.get("/categories", response_model=StandardResponse)
async def get_instruction_categories(
    library_id: int = Query(..., description="指令库ID"),
    db: Session = Depends(get_new_db)
):
    """获取指令分类列表"""
    try:
        # 验证指令库是否存在
        library = db.query(DBInstructionLibraryMaster).filter(
            DBInstructionLibraryMaster.id == library_id
        ).first()
        
        if not library:
            return error_response(msg="指令库不存在", code=ErrorCodes.NOT_FOUND)
        
        categories = db.query(DBInstructionData.category).filter(
            DBInstructionData.library_id == library_id,
            DBInstructionData.category.isnot(None),
            DBInstructionData.category != ''
        ).distinct().all()
        
        category_list = [cat[0] for cat in categories if cat[0]]
        
        return success_response(
            data={
                "library_id": library_id,
                "categories": category_list,
                "count": len(category_list)
            },
            msg=f"获取分类列表成功，共 {len(category_list)} 个分类"
        )
        
    except Exception as e:
        return error_response(msg=f"获取分类列表失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR)


@router.get("/template")
async def download_instruction_template():
    """下载指令导入模板"""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "指令导入模板"
        
        # 定义列标题 - 按照402文件夹实际格式
        headers = [
            "分类", "指令名称", "指令标识", "指令描述", 
            "关联词槽", "追问话术", "追问失败话术", "追问次数",
            "相似问", "执行成功话术", "执行失败话术"
        ]
        
        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        desc_font = Font(italic=True, color="666666")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入标题行
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加示例数据 - 按照402文件夹实际格式
        example_data = [
            [
                "设备控制", "开灯指令", "LIGHT_ON", "打开灯光设备",
                "设备名称,房间位置", "请问您要打开哪个房间的灯？", "抱歉，我没有理解您的意思", "2",
                "开灯|打开灯|亮灯|开启灯光", "灯光已为您打开", "抱歉，灯光打开失败，请稍后重试"
            ],
            [
                "设备控制", "关灯指令", "LIGHT_OFF", "关闭灯光设备",
                "设备名称,房间位置", "请问您要关闭哪个房间的灯？", "抱歉，我没有理解您的意思", "2",
                "关灯|关闭灯|熄灯|关闭灯光", "灯光已为您关闭", "抱歉，灯光关闭失败，请稍后重试"
            ],
            [
                "信息查询", "查询天气", "WEATHER_QUERY", "查询天气信息",
                "地点,日期", "请问您要查询哪个地方的天气？", "抱歉，我没有理解您的意思", "2",
                "天气怎么样|今天天气|天气预报|查天气", "今天天气晴朗，温度适宜", "抱歉，天气查询失败，请稍后重试"
            ]
        ]
        
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # 特殊列宽调整 - 按照实际列调整
        ws.column_dimensions['A'].width = 12  # 分类
        ws.column_dimensions['B'].width = 15  # 指令名称
        ws.column_dimensions['C'].width = 15  # 指令标识
        ws.column_dimensions['D'].width = 20  # 指令描述
        ws.column_dimensions['E'].width = 15  # 关联词槽
        ws.column_dimensions['F'].width = 25  # 追问话术
        ws.column_dimensions['G'].width = 25  # 追问失败话术
        ws.column_dimensions['H'].width = 10  # 追问次数
        ws.column_dimensions['I'].width = 30  # 相似问
        ws.column_dimensions['J'].width = 25  # 执行成功话术
        ws.column_dimensions['K'].width = 25  # 执行失败话术
        
        # 保存到临时文件
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        wb.save(temp_file.name)
        
        return FileResponse(
            temp_file.name,
            filename='instruction_template.xlsx',
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"生成模板失败: {str(e)}")


@router.get("/list", response_model=StandardResponse)
async def get_instruction_list(
    library_id: int = Query(..., description="指令库ID"),
    skip: int = Query(0, ge=0, description="跳过的记录数"),
    limit: int = Query(100, ge=1, le=10000, description="返回的记录数"),
    is_enabled: Optional[bool] = Query(None, description="是否只返回启用的指令"),
    category: Optional[str] = Query(None, description="指令分类筛选"),
    search: Optional[str] = Query(None, description="搜索关键词"),
    db: Session = Depends(get_new_db)
):
    """
    获取指令数据列表
    
    Args:
        library_id: 指令库ID
        skip: 跳过的记录数
        limit: 返回的记录数限制
        is_enabled: 是否只返回启用的指令
        category: 指令分类筛选
        search: 搜索关键词(在指令名称、编码、描述中搜索)
        db: 数据库会话
    
    Returns:
        指令数据列表
    """
    try:
        # 验证指令库是否存在
        library = db.query(DBInstructionLibraryMaster).filter(
            DBInstructionLibraryMaster.id == library_id
        ).first()
        
        if not library:
            return error_response(msg="指令库不存在", code=ErrorCodes.NOT_FOUND)
        
        # 构建查询
        query = db.query(DBInstructionData).filter(
            DBInstructionData.library_id == library_id
        )
        
        # 添加筛选条件
        if is_enabled is not None:
            query = query.filter(DBInstructionData.is_enabled == is_enabled)
        
        if category:
            query = query.filter(DBInstructionData.category == category)
        
        if search:
            search_filter = (
                DBInstructionData.instruction_name.contains(search) |
                DBInstructionData.instruction_code.contains(search) |
                DBInstructionData.instruction_desc.contains(search)
            )
            query = query.filter(search_filter)
        
        # 获取总数（用于分页信息）
        total_count = query.count()
        
        # 排序和分页
        instructions = query.order_by(
            DBInstructionData.sort_order.asc(),
            DBInstructionData.created_time.desc()
        ).offset(skip).limit(limit).all()
        
        # 转换为字典格式（不使用Pydantic模型）
        result = []
        for instruction in instructions:
            # 获取相似问数量
            similar_questions_count = db.query(DBSimilarQuestion).filter(
                DBSimilarQuestion.instruction_id == instruction.id
            ).count()
            
            instruction_dict = {
                "id": instruction.id,
                "library_id": instruction.library_id,
                "instruction_name": instruction.instruction_name,
                "instruction_code": instruction.instruction_code,
                "instruction_desc": instruction.instruction_desc,
                "category": instruction.category,
                "is_slot_related": instruction.is_slot_related,
                "related_slot_ids": instruction.related_slot_ids,
                "success_response": instruction.success_response,
                "failure_response": instruction.failure_response,
                "is_enabled": instruction.is_enabled,
                "sort_order": instruction.sort_order,
                "created_time": instruction.created_time,
                "updated_time": instruction.updated_time,
                "similar_questions_count": similar_questions_count
            }
            result.append(instruction_dict)
        
        return success_response(
            data={
                "instructions": result,
                "total": total_count,
                "skip": skip,
                "limit": limit,
                "has_more": (skip + limit) < total_count
            },
            msg=f"获取指令列表成功，共 {total_count} 条"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        return error_response(msg=f"获取指令列表失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR)


@router.post("/create", response_model=ApiResponse)
async def create_instruction(
    instruction_data: InstructionDataCreate,
    db: Session = Depends(get_new_db)
):
    """
    创建新的指令数据
    
    Args:
        instruction_data: 指令创建数据
        db: 数据库会话
    
    Returns:
        创建结果
    """
    try:
        # 验证指令库是否存在
        library = db.query(DBInstructionLibraryMaster).filter(
            DBInstructionLibraryMaster.id == instruction_data.library_id
        ).first()
        
        if not library:
            raise HTTPException(status_code=404, detail="指令库不存在")
        
        # 检查指令编码是否重复
        existing = db.query(DBInstructionData).filter(
            DBInstructionData.library_id == instruction_data.library_id,
            DBInstructionData.instruction_code == instruction_data.instruction_code
        ).first()
        
        if existing:
            raise HTTPException(status_code=400, detail="指令编码已存在")
        
        # 提取相似问列表
        similar_questions = instruction_data.similar_questions or []
        instruction_dict = instruction_data.dict()
        del instruction_dict['similar_questions']  # 移除相似问，单独处理
        
        # 创建指令数据
        new_instruction = DBInstructionData(**instruction_dict)
        db.add(new_instruction)
        db.flush()  # 获取ID
        
        # 创建相似问
        for i, question_text in enumerate(similar_questions):
            similar_question = DBSimilarQuestion(
                instruction_id=new_instruction.id,
                question_text=question_text,
                sort_order=i + 1,
                is_enabled=True
            )
            db.add(similar_question)
        
        db.commit()
        
        return ApiResponse(
            success=True,
            message=f"指令 '{instruction_data.instruction_name}' 创建成功",
            data={
                "instruction_id": new_instruction.id,
                "similar_questions_count": len(similar_questions)
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"创建指令失败: {str(e)}")


@router.get("/{instruction_id}", response_model=dict)
async def get_instruction_detail(
    instruction_id: int,
    db: Session = Depends(get_new_db)
):
    """
    获取指令详情(包含相似问)
    
    Args:
        instruction_id: 指令ID
        db: 数据库会话
    
    Returns:
        指令详情信息
    """
    try:
        instruction = db.query(DBInstructionData).filter(
            DBInstructionData.id == instruction_id
        ).first()
        
        if not instruction:
            raise HTTPException(status_code=404, detail="指令不存在")
        
        # 获取相似问
        similar_questions = db.query(DBSimilarQuestion).filter(
            DBSimilarQuestion.instruction_id == instruction_id
        ).order_by(DBSimilarQuestion.sort_order.asc()).all()
        
        # 构建返回数据
        result = {
            "instruction": instruction,
            "similar_questions": similar_questions,
            "similar_questions_count": len(similar_questions)
        }
        
        return result
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取指令详情失败: {str(e)}")


@router.put("/{instruction_id}", response_model=ApiResponse)
async def update_instruction(
    instruction_id: int,
    instruction_data: dict,
    db: Session = Depends(get_new_db)
):
    """
    更新指令数据
    
    Args:
        instruction_id: 指令ID
        instruction_data: 更新数据
        db: 数据库会话
    
    Returns:
        更新结果
    """
    try:
        instruction = db.query(DBInstructionData).filter(
            DBInstructionData.id == instruction_id
        ).first()
        
        if not instruction:
            raise HTTPException(status_code=404, detail="指令不存在")
        
        # 检查指令编码重复(如果要更新编码)
        if "instruction_code" in instruction_data:
            if instruction_data["instruction_code"] != instruction.instruction_code:
                existing = db.query(DBInstructionData).filter(
                    DBInstructionData.library_id == instruction.library_id,
                    DBInstructionData.instruction_code == instruction_data["instruction_code"],
                    DBInstructionData.id != instruction_id
                ).first()
                
                if existing:
                    raise HTTPException(status_code=400, detail="指令编码已存在")
        
        # 更新字段
        for field, value in instruction_data.items():
            if hasattr(instruction, field) and field != 'similar_questions':
                setattr(instruction, field, value)
        
        db.commit()
        
        return ApiResponse(
            success=True,
            message=f"指令 '{instruction.instruction_name}' 更新成功",
            data={"instruction_id": instruction_id}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"更新指令失败: {str(e)}")


@router.delete("/{instruction_id}", response_model=ApiResponse)
async def delete_instruction(
    instruction_id: int,
    db: Session = Depends(get_new_db)
):
    """
    删除指令数据(级联删除相似问)
    
    Args:
        instruction_id: 指令ID
        db: 数据库会话
    
    Returns:
        删除结果
    """
    try:
        instruction = db.query(DBInstructionData).filter(
            DBInstructionData.id == instruction_id
        ).first()
        
        if not instruction:
            raise HTTPException(status_code=404, detail="指令不存在")
        
        instruction_name = instruction.instruction_name
        
        # 删除指令(级联删除相似问)
        db.delete(instruction)
        db.commit()
        
        return ApiResponse(
            success=True,
            message=f"指令 '{instruction_name}' 删除成功",
            data={"deleted_instruction_id": instruction_id}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"删除指令失败: {str(e)}")


# ===== 相似问管理 =====

@router.get("/{instruction_id}/similar-questions", response_model=StandardResponse)
async def get_similar_questions(
    instruction_id: int,
    db: Session = Depends(get_new_db)
):
    """获取指令的相似问列表"""
    try:
        # 验证指令是否存在
        instruction = db.query(DBInstructionData).filter(
            DBInstructionData.id == instruction_id
        ).first()
        
        if not instruction:
            return error_response(msg="指令不存在", code=ErrorCodes.NOT_FOUND)
        
        similar_questions = db.query(DBSimilarQuestion).filter(
            DBSimilarQuestion.instruction_id == instruction_id
        ).order_by(DBSimilarQuestion.sort_order.asc()).all()
        
        # 转换为字典格式
        result = []
        for sq in similar_questions:
            sq_dict = {
                "id": sq.id,
                "instruction_id": sq.instruction_id,
                "question_text": sq.question_text,
                "is_enabled": sq.is_enabled,
                "sort_order": sq.sort_order,
                "created_time": sq.created_time
            }
            result.append(sq_dict)
        
        return success_response(
            data={
                "similar_questions": result,
                "total": len(result),
                "instruction_id": instruction_id,
                "instruction_name": instruction.instruction_name
            },
            msg=f"获取相似问列表成功，共 {len(result)} 条"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        return error_response(msg=f"获取相似问失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR)


@router.post("/{instruction_id}/similar-questions", response_model=ApiResponse)
async def add_similar_question(
    instruction_id: int,
    question_data: SimilarQuestionCreate,
    db: Session = Depends(get_new_db)
):
    """添加相似问"""
    try:
        # 验证指令是否存在
        instruction = db.query(DBInstructionData).filter(
            DBInstructionData.id == instruction_id
        ).first()
        
        if not instruction:
            raise HTTPException(status_code=404, detail="指令不存在")
        
        # 检查相似问是否重复
        existing = db.query(DBSimilarQuestion).filter(
            DBSimilarQuestion.instruction_id == instruction_id,
            DBSimilarQuestion.question_text == question_data.question_text
        ).first()
        
        if existing:
            raise HTTPException(status_code=400, detail="相似问已存在")
        
        # 创建相似问
        question_dict = question_data.dict()
        question_dict['instruction_id'] = instruction_id
        
        new_question = DBSimilarQuestion(**question_dict)
        db.add(new_question)
        db.commit()
        
        return ApiResponse(
            success=True,
            message="相似问添加成功",
            data={"question_id": new_question.id}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"添加相似问失败: {str(e)}")


@router.put("/{instruction_id}/similar-questions/{question_id}", response_model=ApiResponse)
async def update_similar_question(
    instruction_id: int,
    question_id: int,
    question_data: dict,
    db: Session = Depends(get_new_db)
):
    """更新相似问"""
    try:
        question = db.query(DBSimilarQuestion).filter(
            DBSimilarQuestion.id == question_id,
            DBSimilarQuestion.instruction_id == instruction_id
        ).first()
        
        if not question:
            raise HTTPException(status_code=404, detail="相似问不存在")
        
        # 检查文本重复(如果要更新文本)
        if "question_text" in question_data:
            if question_data["question_text"] != question.question_text:
                existing = db.query(DBSimilarQuestion).filter(
                    DBSimilarQuestion.instruction_id == instruction_id,
                    DBSimilarQuestion.question_text == question_data["question_text"],
                    DBSimilarQuestion.id != question_id
                ).first()
                
                if existing:
                    raise HTTPException(status_code=400, detail="相似问已存在")
        
        # 更新字段
        for field, value in question_data.items():
            if hasattr(question, field):
                setattr(question, field, value)
        
        db.commit()
        
        return ApiResponse(
            success=True,
            message="相似问更新成功",
            data={"question_id": question_id}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"更新相似问失败: {str(e)}")


@router.delete("/{instruction_id}/similar-questions/{question_id}", response_model=ApiResponse)
async def delete_similar_question(
    instruction_id: int,
    question_id: int,
    db: Session = Depends(get_new_db)
):
    """删除相似问"""
    try:
        question = db.query(DBSimilarQuestion).filter(
            DBSimilarQuestion.id == question_id,
            DBSimilarQuestion.instruction_id == instruction_id
        ).first()
        
        if not question:
            raise HTTPException(status_code=404, detail="相似问不存在")
        
        db.delete(question)
        db.commit()
        
        return ApiResponse(
            success=True,
            message="相似问删除成功",
            data={"deleted_question_id": question_id}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"删除相似问失败: {str(e)}")


# ===== 批量操作 =====

@router.post("/batch-import", response_model=StandardResponse)
async def batch_import_instructions(
    library_id: int = Form(..., description="指令库ID"),
    file: UploadFile = File(..., description="Excel文件"),
    db: Session = Depends(get_new_db)
):
    """
    批量导入指令数据
    
    Args:
        library_id: 指令库ID
        file: Excel文件
        db: 数据库会话
    
    Returns:
        导入结果
    """
    try:
        # 验证指令库是否存在
        library = db.query(DBInstructionLibraryMaster).filter(
            DBInstructionLibraryMaster.id == library_id
        ).first()
        
        if not library:
            raise HTTPException(status_code=404, detail="指令库不存在")
        
        # 验证文件类型
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="文件格式错误，请上传Excel文件")
        
        # 读取Excel文件
        content = await file.read()
        df = pd.read_excel(BytesIO(content))
        
        # 验证必需的列 - 按照402文件夹格式
        required_columns = ['指令名称', '指令标识']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            raise HTTPException(
                status_code=400, 
                detail=f"Excel文件缺少必需的列: {', '.join(missing_columns)}"
            )
        
        success_count = 0
        failed_count = 0
        errors = []
        imported_ids = []
        
        # 批量导入 - 按照402文件夹格式
        for index, row in df.iterrows():
            try:
                # 安全地获取指令名称和标识，跳过空行
                instruction_name = str(row.get('指令名称', '')).strip() if pd.notna(row.get('指令名称')) else ''
                instruction_code = str(row.get('指令标识', '')).strip() if pd.notna(row.get('指令标识')) else ''
                
                # 跳过没有指令名称或标识的行（通常是相似问行）
                if not instruction_name or not instruction_code:
                    continue
                
                # 检查指令编码是否重复
                existing = db.query(DBInstructionData).filter(
                    DBInstructionData.library_id == library_id,
                    DBInstructionData.instruction_code == instruction_code
                ).first()
                
                if existing:
                    # 如果指令已存在，跳过指令创建，但记录ID用于后续相似问处理
                    imported_ids.append(existing.id)
                    continue
                
                # 安全地获取其他字段
                instruction_desc = str(row.get('指令描述', '')).strip() if pd.notna(row.get('指令描述')) else ''
                category = str(row.get('分类', '')).strip() if pd.notna(row.get('分类')) else ''
                
                # 检查关联词槽格式
                related_slot_ids = ""
                if '关联词槽' in row and pd.notna(row['关联词槽']):
                    # 402格式中关联词槽是用逗号分隔的词槽名称
                    related_slot_ids = str(row['关联词槽']).strip()
                
                # 智能解析：如果指令名称包含+号，尝试推断多个词槽
                if '+' in instruction_name:
                    # 从指令名称中提取可能的词槽名称
                    name_parts = instruction_name.split('+')
                    potential_slots = []
                    
                    # 已有的关联词槽
                    if related_slot_ids:
                        potential_slots.append(related_slot_ids)
                    
                    # 从指令名称中推断的词槽
                    for part in name_parts:
                        part = part.strip()
                        # 移除常见的动词前缀
                        for prefix in ['设置', '调节', '选择', '播放', '打开', '关闭']:
                            if part.startswith(prefix):
                                part = part[len(prefix):]
                                break
                        
                        if part and part not in potential_slots:
                            potential_slots.append(part)
                    
                    # 合并所有词槽
                    if len(potential_slots) > 1:
                        related_slot_ids = ','.join(potential_slots)
                
                # 安全地获取话术
                success_response_text = str(row.get('执行成功话术', '')).strip() if pd.notna(row.get('执行成功话术')) else ''
                failure_response_text = str(row.get('执行失败话术', '')).strip() if pd.notna(row.get('执行失败话术')) else ''
                
                # 创建指令数据 - 映射402格式到数据库字段
                instruction_data = {
                    'library_id': library_id,
                    'instruction_name': instruction_name,
                    'instruction_code': instruction_code,
                    'instruction_desc': instruction_desc,
                    'category': category,
                    'is_slot_related': bool(related_slot_ids),  # 根据是否有关联词槽判断
                    'related_slot_ids': related_slot_ids,
                    'success_response': success_response_text,
                    'failure_response': failure_response_text,
                    'is_enabled': True,  # 默认启用
                    'sort_order': success_count + 1  # 按成功导入顺序排序
                }
                
                new_instruction = DBInstructionData(**instruction_data)
                db.add(new_instruction)
                db.flush()
                
                # 暂时不处理相似问，稍后统一处理
                # 402格式中相似问是分散在不同行中的
                
                imported_ids.append(new_instruction.id)
                success_count += 1
                
            except Exception as e:
                errors.append(f"第{index+2}行: {str(e)}")
                failed_count += 1
        
        # 处理相似问数据 - 根据402文件格式
        similar_success_count = 0
        
        # 重新遍历数据，处理相似问
        if success_count > 0 or imported_ids:
            # 找到所有指令行的索引
            instruction_indices = []
            for index, row in df.iterrows():
                instruction_name = str(row.get('指令名称', '')).strip() if pd.notna(row.get('指令名称')) else ''
                instruction_code = str(row.get('指令标识', '')).strip() if pd.notna(row.get('指令标识')) else ''
                
                if instruction_name and instruction_code:
                    instruction_indices.append(index)
            
            # 处理每个指令的相似问
            for i, instruction_index in enumerate(instruction_indices):
                try:
                    # 获取指令信息
                    instruction_row = df.iloc[instruction_index]
                    instruction_code = str(instruction_row.get('指令标识', '')).strip()
                    
                    # 找到对应的数据库记录（包括新创建的和已存在的）
                    instruction_record = db.query(DBInstructionData).filter(
                        DBInstructionData.library_id == library_id,
                        DBInstructionData.instruction_code == instruction_code
                    ).first()
                    
                    if not instruction_record:
                        continue
                    
                    # 清空现有相似问（如果是重新导入）
                    db.query(DBSimilarQuestion).filter(
                        DBSimilarQuestion.instruction_id == instruction_record.id
                    ).delete()
                    
                    # 确定相似问的范围（从当前指令行+1到下一个指令行-1）
                    start_index = instruction_index + 1
                    end_index = instruction_indices[i + 1] if i + 1 < len(instruction_indices) else len(df)
                    
                    # 处理相似问行
                    for similar_index in range(start_index, end_index):
                        if similar_index >= len(df):
                            break
                            
                        similar_row = df.iloc[similar_index]
                        similar_question_text = str(similar_row.get('相似问', '')).strip() if pd.notna(similar_row.get('相似问')) else ''
                        
                        if similar_question_text:
                            # 检查相似问是否已存在
                            existing_similar = db.query(DBSimilarQuestion).filter(
                                DBSimilarQuestion.instruction_id == instruction_record.id,
                                DBSimilarQuestion.question_text == similar_question_text
                            ).first()
                            
                            if not existing_similar:
                                # 创建相似问
                                similar_question = DBSimilarQuestion(
                                    instruction_id=instruction_record.id,
                                    question_text=similar_question_text,
                                    is_enabled=True,
                                    sort_order=similar_success_count + 1
                                )
                                db.add(similar_question)
                                similar_success_count += 1
                
                except Exception as e:
                    # 相似问导入失败不影响整体流程
                    continue
        
        # 提交事务
        if success_count > 0 or similar_success_count > 0:
            db.commit()
        else:
            db.rollback()
        
        return success_response(
            data={
                "success_count": success_count,
                "failed_count": failed_count,
                "similar_questions_count": similar_success_count,
                "errors": errors[:10],  # 只返回前10个错误
                "imported_ids": imported_ids
            },
            msg=f"批量导入完成：指令 {success_count} 条，相似问 {similar_success_count} 条，失败 {failed_count} 条"
        )
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return error_response(msg=f"批量导入失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR)


@router.get("/stats/{library_id}")
async def get_instruction_stats(
    library_id: int,
    db: Session = Depends(get_new_db)
):
    """获取指令库的指令统计信息"""
    try:
        # 验证指令库是否存在
        library = db.query(DBInstructionLibraryMaster).filter(
            DBInstructionLibraryMaster.id == library_id
        ).first()
        
        if not library:
            raise HTTPException(status_code=404, detail="指令库不存在")
        
        # 统计信息
        total_instructions = db.query(DBInstructionData).filter(
            DBInstructionData.library_id == library_id
        ).count()
        
        enabled_instructions = db.query(DBInstructionData).filter(
            DBInstructionData.library_id == library_id,
            DBInstructionData.is_enabled == True
        ).count()
        
        # 按分类统计
        categories = db.query(
            DBInstructionData.category,
            db.func.count(DBInstructionData.id).label('count')
        ).filter(
            DBInstructionData.library_id == library_id
        ).group_by(DBInstructionData.category).all()
        
        category_stats = [{"category": cat or "未分类", "count": count} for cat, count in categories]
        
        # 相似问统计
        total_similar_questions = db.query(DBSimilarQuestion).join(
            DBInstructionData, DBSimilarQuestion.instruction_id == DBInstructionData.id
        ).filter(
            DBInstructionData.library_id == library_id
        ).count()
        
        return {
            "library_id": library_id,
            "library_name": library.name,
            "total_instructions": total_instructions,
            "enabled_instructions": enabled_instructions,
            "disabled_instructions": total_instructions - enabled_instructions,
            "total_similar_questions": total_similar_questions,
            "categories": category_stats
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取统计信息失败: {str(e)}")


