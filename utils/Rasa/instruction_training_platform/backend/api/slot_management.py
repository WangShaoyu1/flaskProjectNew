"""
智能对话训练平台 - 词槽管理API
提供词槽定义和词槽值的CRUD操作，支持指令与词槽关联管理
"""

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
import json
import pandas as pd
from io import BytesIO
import sys
import os
import tempfile
sys.path.append(os.path.dirname(os.path.dirname(__file__)))

from models.database_models import (
    get_new_db, 
    SlotDefinition as DBSlotDefinition,
    SlotValue as DBSlotValue,
    InstructionLibraryMaster as DBInstructionLibraryMaster,
    InstructionData as DBInstructionData
)
from response_utils import StandardResponse, success_response, error_response, ErrorCodes, Messages
from models.schemas import (
    SlotDefinition, SlotDefinitionCreate, 
    SlotValue, SlotValueBase,
    ApiResponse
)

router = APIRouter(prefix="/api/v2/slots", tags=["词槽管理"])

# ===== 辅助功能(必须放在动态路由之前) =====

@router.get("/slot-types")
async def get_slot_types():
    """获取支持的词槽类型"""
    return success_response(data={"slot_types": [
            {"value": "categorical", "label": "分类型", "description": "有限的预定义值"},
            {"value": "text", "label": "文本型", "description": "任意文本内容"},
            {"value": "float", "label": "数值型", "description": "数字值"},
            {"value": "boolean", "label": "布尔型", "description": "真/假值"}
        ]
    })


@router.post("/init-system-slots", response_model=StandardResponse)
async def init_system_slots(
    library_id: int = Query(..., description="指令库ID"),
    db: Session = Depends(get_new_db)
):
    """为指定指令库初始化系统词槽"""
    try:
        # 验证指令库是否存在
        library = db.query(DBInstructionLibraryMaster).filter(
            DBInstructionLibraryMaster.id == library_id
        ).first()
        
        if not library:
            return error_response(msg="指令库不存在", code=ErrorCodes.NOT_FOUND)
        
        # 系统词槽定义
        SYSTEM_SLOTS = [
            {
                "slot_name": "时长",
                "slot_name_en": "duration",
                "slot_type": "text",
                "description": "时间持续长度，如：15分钟、2小时、3天等",
                "is_system": True,
                "is_required": False,
                "is_active": True
            },
            {
                "slot_name": "数字",
                "slot_name_en": "number",
                "slot_type": "float",
                "description": "数值类型，如：1、2.5、100等",
                "is_system": True,
                "is_required": False,
                "is_active": True
            },
            {
                "slot_name": "数量",
                "slot_name_en": "quantity",
                "slot_type": "text",
                "description": "数量表达，如：一个、两台、三件等",
                "is_system": True,
                "is_required": False,
                "is_active": True
            },
            {
                "slot_name": "时间",
                "slot_name_en": "time",
                "slot_type": "text",
                "description": "时间点，如：上午9点、下午3点、晚上8点等",
                "is_system": True,
                "is_required": False,
                "is_active": True
            },
            {
                "slot_name": "音量",
                "slot_name_en": "volume",
                "slot_type": "text",
                "description": "音量大小，如：大声、小声、50%等",
                "is_system": True,
                "is_required": False,
                "is_active": True
            }
        ]
        
        created_count = 0
        updated_count = 0
        
        for slot_config in SYSTEM_SLOTS:
            # 检查系统词槽是否已存在
            existing_slot = db.query(DBSlotDefinition).filter(
                DBSlotDefinition.library_id == library_id,
                DBSlotDefinition.slot_name_en == slot_config["slot_name_en"],
                DBSlotDefinition.is_system == True
            ).first()
            
            if existing_slot:
                # 更新现有系统词槽
                for key, value in slot_config.items():
                    if key != "slot_name_en":  # 不更新英文名
                        setattr(existing_slot, key, value)
                updated_count += 1
            else:
                # 创建新的系统词槽
                slot_data = {
                    "library_id": library_id,
                    **slot_config
                }
                new_slot = DBSlotDefinition(**slot_data)
                db.add(new_slot)
                created_count += 1
        
        # 提交事务
        db.commit()
        
        return success_response(
            data={
                "created_count": created_count,
                "updated_count": updated_count,
                "total_system_slots": len(SYSTEM_SLOTS)
            },
            msg=f"系统词槽初始化完成：新增 {created_count} 个，更新 {updated_count} 个"
        )
        
    except Exception as e:
        db.rollback()
        return error_response(msg=f"初始化系统词槽失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR)


@router.get("/download-template")
async def download_slot_template():
    """下载词槽导入模板"""
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        
        # 创建工作簿
        wb = Workbook()
        
        # 修改默认工作表
        ws = wb.active
        ws.title = "词槽"
        
        # 词槽列标题 - 按照402文件夹实际格式
        headers = [
            "词槽名称", "词槽描述", "实体ID", "实体标准名", "实体别名"
        ]
        
        # 设置样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        desc_font = Font(italic=True, color="666666")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入标题行
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加示例数据 - 按照402文件夹实际格式
        example_data = [
            ["休眠时间", "", 1, "15秒", "十五秒"],
            ["", "", 2, "30秒", "三十秒"],
            ["", "", 3, "1分钟", "1分==一分钟==一分"],
            ["", "", 4, "2分钟", "2分==两分钟==二分钟==两分"],
            ["", "", 5, "5分钟", "5分==五分钟==五分"],
            ["设备名称", "智能家居设备名称", 6, "灯", "灯光==电灯==照明灯"],
            ["", "", 7, "空调", "空调机==冷气==AC"],
            ["", "", 8, "电视", "电视机==TV==电视机"],
            ["房间位置", "房间或区域位置", 9, "客厅", "大厅==起居室==客厅"],
            ["", "", 10, "卧室", "卧房==睡房==房间"],
            ["", "", 11, "厨房", "厨房==烹饪间"]
        ]
        
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15  # 词槽名称
        ws.column_dimensions['B'].width = 20  # 词槽描述
        ws.column_dimensions['C'].width = 10  # 实体ID
        ws.column_dimensions['D'].width = 15  # 实体标准名
        ws.column_dimensions['E'].width = 30  # 实体别名
        
        # 保存到临时文件
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_file.close()
        wb.save(temp_file.name)
        
        return FileResponse(
            temp_file.name,
            filename='slot_template.xlsx',
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"生成模板失败: {str(e)}")


@router.post("/batch-import", response_model=StandardResponse)
async def batch_import_slots(
    library_id: int = Form(..., description="指令库ID"),
    file: UploadFile = File(..., description="Excel文件"),
    db: Session = Depends(get_new_db)
):
    """批量导入词槽"""
    try:
        # 验证指令库是否存在
        library = db.query(DBInstructionLibraryMaster).filter(
            DBInstructionLibraryMaster.id == library_id
        ).first()
        
        if not library:
            return error_response(msg="指令库不存在", code=ErrorCodes.NOT_FOUND)
        
        # 验证文件类型
        if not file.filename.endswith(('.xlsx', '.xls')):
            return error_response(msg="文件格式错误，请上传Excel文件", code=ErrorCodes.PARAM_ERROR)
        
        # 读取Excel文件
        content = await file.read()
        
        # 读取Excel文件 - 按照402格式(单个工作表)
        slots_df = pd.read_excel(BytesIO(content))
        
        # 验证必需的列 - 按照402文件夹格式
        required_columns = ['词槽名称', '实体标准名']
        missing_columns = [col for col in required_columns if col not in slots_df.columns]
        
        if missing_columns:
            return error_response(
                msg=f"Excel文件缺少必需的列: {', '.join(missing_columns)}",
                code=ErrorCodes.PARAM_ERROR
            )
        
        success_count = 0
        failed_count = 0
        errors = []
        imported_slot_ids = []
        
        # 批量导入词槽定义和词槽值 - 按照402格式处理
        slot_definitions = {}  # 用于存储词槽定义
        slot_values = {}       # 用于存储词槽值
        
        # 第一步：收集数据
        for index, row in slots_df.iterrows():
            try:
                # 安全地获取字符串值，处理NaN情况
                slot_name = str(row.get('词槽名称', '')).strip() if pd.notna(row.get('词槽名称')) else ''
                slot_desc = str(row.get('词槽描述', '')).strip() if pd.notna(row.get('词槽描述')) else ''
                entity_id = row.get('实体ID', '') if pd.notna(row.get('实体ID')) else ''
                standard_value = str(row.get('实体标准名', '')).strip() if pd.notna(row.get('实体标准名')) else ''
                synonyms = str(row.get('实体别名', '')).strip() if pd.notna(row.get('实体别名')) else ''
                
                # 如果有词槽名称，则这是一个新的词槽定义
                if slot_name:
                    if slot_name not in slot_definitions:
                        slot_definitions[slot_name] = {
                            'slot_name': slot_name,
                            'slot_desc': slot_desc,
                            'slot_name_en': slot_name.upper().replace(' ', '_'),  # 生成英文名
                            'slot_type': 'categorical',  # 402格式默认为分类型
                            'values': []
                        }
                    
                    current_slot = slot_name
                else:
                    # 如果没有词槽名称，使用最近的词槽名称
                    if not slot_definitions:
                        errors.append(f"第{index+2}行: 缺少词槽名称")
                        failed_count += 1
                        continue
                    current_slot = list(slot_definitions.keys())[-1]
                
                # 添加词槽值
                if standard_value:
                    slot_definitions[current_slot]['values'].append({
                        'standard_value': standard_value,
                        'synonyms': synonyms,
                        'entity_id': entity_id
                    })
                
            except Exception as e:
                errors.append(f"第{index+2}行: 数据解析错误 - {str(e)}")
                failed_count += 1
        
        # 第二步：导入词槽定义和词槽值
        for slot_name, slot_info in slot_definitions.items():
            try:
                # 检查词槽是否已存在（优先按中文名匹配）
                existing = db.query(DBSlotDefinition).filter(
                    DBSlotDefinition.library_id == library_id,
                    DBSlotDefinition.slot_name == slot_name
                ).first()
                
                if not existing:
                    # 如果按中文名没找到，再按英文名查找
                    existing = db.query(DBSlotDefinition).filter(
                        DBSlotDefinition.library_id == library_id,
                        DBSlotDefinition.slot_name_en == slot_info['slot_name_en']
                    ).first()
                
                if existing:
                    # 词槽已存在，更新词槽值
                    print(f"词槽 '{slot_name}' 已存在，更新词槽值...")
                    
                    # 清空现有词槽值
                    db.query(DBSlotValue).filter(
                        DBSlotValue.slot_id == existing.id
                    ).delete()
                    
                    # 添加新的词槽值
                    for i, value_info in enumerate(slot_info['values']):
                        slot_value_data = {
                            'slot_id': existing.id,
                            'standard_value': value_info['standard_value'],
                            'aliases': value_info['synonyms'],
                            'is_active': True,
                            'sort_order': i + 1
                        }
                        
                        new_value = DBSlotValue(**slot_value_data)
                        db.add(new_value)
                    
                    imported_slot_ids.append(existing.id)
                    success_count += 1
                    
                else:
                    # 词槽不存在，创建新词槽
                    slot_data = {
                        'library_id': library_id,
                        'slot_name': slot_info['slot_name'],
                        'slot_name_en': slot_info['slot_name_en'],
                        'slot_type': slot_info['slot_type'],
                        'description': slot_info['slot_desc'],
                        'is_required': False,
                        'is_active': True
                    }
                    
                    new_slot = DBSlotDefinition(**slot_data)
                    db.add(new_slot)
                    db.flush()
                    
                    # 添加词槽值
                    for i, value_info in enumerate(slot_info['values']):
                        slot_value_data = {
                            'slot_id': new_slot.id,
                            'standard_value': value_info['standard_value'],
                            'aliases': value_info['synonyms'],
                            'is_active': True,
                            'sort_order': i + 1
                        }
                        
                        new_value = DBSlotValue(**slot_value_data)
                        db.add(new_value)
                    
                    imported_slot_ids.append(new_slot.id)
                    success_count += 1
                
            except Exception as e:
                errors.append(f"词槽 '{slot_name}': {str(e)}")
                failed_count += 1
        
        # 提交事务
        if success_count > 0:
            db.commit()
        else:
            db.rollback()
        
        # 统计词槽值数量
        total_values_count = sum(len(slot_info['values']) for slot_info in slot_definitions.values())
        
        return success_response(
            data={
                "slots_success_count": success_count,
                "slots_failed_count": failed_count,
                "total_values_count": total_values_count,
                "errors": errors[:20],  # 只返回前20个错误
                "imported_slot_ids": imported_slot_ids
            },
            msg=f"批量导入完成：词槽 {success_count} 个，词槽值约 {total_values_count} 个，失败 {failed_count} 条"
        )
        
    except Exception as e:
        db.rollback()
        return error_response(msg=f"批量导入失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR)


# ===== 词槽定义管理 =====

@router.get("/list", response_model=StandardResponse)
async def get_slot_list(
    library_id: int = Query(..., description="指令库ID"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    page: Optional[int] = Query(None, ge=1, description="页码(从1开始)"),
    size: Optional[int] = Query(None, ge=1, le=1000, description="每页大小"),
    is_active: Optional[bool] = Query(None),
    slot_type: Optional[str] = Query(None),
    is_system: Optional[bool] = Query(None, description="是否系统词槽"),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_new_db)
):
    """获取词槽定义列表"""
    try:
        # 参数转换：支持page/size和skip/limit两种格式
        if page is not None and size is not None:
            skip = (page - 1) * size
            limit = size
        
        # 验证指令库是否存在
        library = db.query(DBInstructionLibraryMaster).filter(
            DBInstructionLibraryMaster.id == library_id
        ).first()
        
        if not library:
            return error_response(msg="指令库不存在", code=ErrorCodes.NOT_FOUND)
        
        # 构建查询
        query = db.query(DBSlotDefinition).filter(
            DBSlotDefinition.library_id == library_id
        )
        
        # 添加筛选条件
        if is_active is not None:
            query = query.filter(DBSlotDefinition.is_active == is_active)
        
        if slot_type:
            query = query.filter(DBSlotDefinition.slot_type == slot_type)
        
        if is_system is not None:
            query = query.filter(DBSlotDefinition.is_system == is_system)
        
        if search:
            search_filter = (
                DBSlotDefinition.slot_name.contains(search) |
                DBSlotDefinition.slot_name_en.contains(search) |
                DBSlotDefinition.description.contains(search)
            )
            query = query.filter(search_filter)
        
        # 排序和分页 - 自定义词槽优先，系统词槽在后
        slots = query.order_by(
            DBSlotDefinition.is_system.asc(),
            DBSlotDefinition.created_time.desc()
        ).offset(skip).limit(limit).all()
        
        # 转换为Pydantic模型并添加统计信息
        result = []
        for slot in slots:
            # 统计词槽值数量
            values_count = db.query(DBSlotValue).filter(
                DBSlotValue.slot_id == slot.id
            ).count()
            
            # 统计使用此词槽的指令数量
            related_instructions = db.query(DBInstructionData).filter(
                DBInstructionData.related_slot_ids.contains(f'[{slot.id}]') |
                DBInstructionData.related_slot_ids.contains(f'{slot.id}')
            ).count()
            
            slot_dict = {
                "id": slot.id,
                "library_id": slot.library_id,
                "slot_name": slot.slot_name,
                "slot_name_en": slot.slot_name_en,
                "slot_type": slot.slot_type,
                "description": slot.description,
                "is_required": slot.is_required,
                "is_active": slot.is_active,
                "is_system": getattr(slot, 'is_system', False),  # 兼容旧数据
                "created_time": slot.created_time,
                "updated_time": slot.updated_time,
                "values_count": values_count,
                "related_instructions_count": related_instructions
            }
            result.append(SlotDefinition(**slot_dict))
        
        # 获取总数
        total_count = query.count()
        
        return success_response(data={
            "slots": result,
            "total": total_count
        })
    except HTTPException:
        raise
    except Exception as e:
        return error_response(msg=f"获取词槽列表失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR)


@router.post("/create", response_model=StandardResponse)
async def create_slot(
    slot_data: SlotDefinitionCreate,
    db: Session = Depends(get_new_db)
):
    """创建词槽"""
    try:
        # 验证指令库是否存在
        library = db.query(DBInstructionLibraryMaster).filter(
            DBInstructionLibraryMaster.id == slot_data.library_id
        ).first()
        
        if not library:
            return error_response(msg="指令库不存在", code=ErrorCodes.NOT_FOUND)
        
        # 检查词槽英文名是否已存在
        existing_slot = db.query(DBSlotDefinition).filter(
            DBSlotDefinition.library_id == slot_data.library_id,
            DBSlotDefinition.slot_name_en == slot_data.slot_name_en
        ).first()
        
        if existing_slot:
            return error_response(msg="词槽英文名已存在", code=ErrorCodes.PARAM_ERROR)
        
        # 创建词槽
        new_slot = DBSlotDefinition(**slot_data.dict())
        db.add(new_slot)
        db.commit()
        db.refresh(new_slot)
        
        return success_response(data={"slot": new_slot}, msg="词槽创建成功")
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return error_response(msg=f"创建词槽失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR)


@router.put("/{slot_id}", response_model=StandardResponse)
async def update_slot(
    slot_id: int,
    slot_data: SlotDefinitionCreate,
    db: Session = Depends(get_new_db)
):
    """更新词槽"""
    try:
        # 获取词槽
        slot = db.query(DBSlotDefinition).filter(
            DBSlotDefinition.id == slot_id
        ).first()
        
        if not slot:
            return error_response(msg="词槽不存在", code=ErrorCodes.NOT_FOUND)
        
        # 检查词槽英文名是否与其他词槽冲突
        existing_slot = db.query(DBSlotDefinition).filter(
            DBSlotDefinition.library_id == slot_data.library_id,
            DBSlotDefinition.slot_name_en == slot_data.slot_name_en,
            DBSlotDefinition.id != slot_id
        ).first()
        
        if existing_slot:
            return error_response(msg="词槽英文名已存在", code=ErrorCodes.PARAM_ERROR)
        
        # 更新词槽
        for field, value in slot_data.dict().items():
            setattr(slot, field, value)
        
        db.commit()
        db.refresh(slot)
        
        return success_response(data={"slot": slot}, msg="词槽更新成功")
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return error_response(msg=f"更新词槽失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR)


@router.delete("/{slot_id}", response_model=StandardResponse)
async def delete_slot(
    slot_id: int,
    db: Session = Depends(get_new_db)
):
    """删除词槽"""
    try:
        # 获取词槽
        slot = db.query(DBSlotDefinition).filter(
            DBSlotDefinition.id == slot_id
        ).first()
        
        if not slot:
            return error_response(msg="词槽不存在", code=ErrorCodes.NOT_FOUND)
        
        # 检查是否有关联的指令
        related_instructions = db.query(DBInstructionData).filter(
            DBInstructionData.related_slot_ids.contains(f'[{slot_id}]') |
            DBInstructionData.related_slot_ids.contains(f'{slot_id}')
        ).count()
        
        if related_instructions > 0:
            return error_response(msg=f"词槽被 {related_instructions} 个指令使用，无法删除", code=ErrorCodes.PARAM_ERROR)
        
        # 删除词槽（会级联删除词槽值）
        db.delete(slot)
        db.commit()
        
        return success_response(msg="词槽删除成功")
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return error_response(msg=f"删除词槽失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR)


# ===== 动态路由放在最后 =====

@router.get("/{slot_id}")
async def get_slot_detail(
    slot_id: int,
    db: Session = Depends(get_new_db)
):
    """获取词槽详情(包含词槽值)"""
    try:
        slot = db.query(DBSlotDefinition).filter(
            DBSlotDefinition.id == slot_id
        ).first()
        
        if not slot:
            return error_response(msg="词槽不存在", code=ErrorCodes.NOT_FOUND)
        
        # 获取词槽值
        slot_values = db.query(DBSlotValue).filter(
            DBSlotValue.slot_id == slot_id
        ).order_by(DBSlotValue.sort_order.asc()).all()
        
        # 获取使用此词槽的指令数量
        related_instructions = db.query(DBInstructionData).filter(
            DBInstructionData.related_slot_ids.contains(f'[{slot_id}]') |
            DBInstructionData.related_slot_ids.contains(f'{slot_id}')
        ).count()
        
        # 构建返回数据
        result = {
            "slot": slot,
            "slot_values": slot_values,
            "values_count": len(slot_values),
            "related_instructions_count": related_instructions
        }
        
        return success_response(data=result)
    except HTTPException:
        raise
    except Exception as e:
        return error_response(msg=f"获取词槽详情失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR)


@router.get("/{slot_id}/values", response_model=StandardResponse)
async def get_slot_values(
    slot_id: int,
    db: Session = Depends(get_new_db)
):
    """获取词槽值列表"""
    try:
        # 验证词槽是否存在
        slot = db.query(DBSlotDefinition).filter(
            DBSlotDefinition.id == slot_id
        ).first()
        
        if not slot:
            return error_response(msg="词槽不存在", code=ErrorCodes.NOT_FOUND)
        
        # 获取词槽值
        slot_values = db.query(DBSlotValue).filter(
            DBSlotValue.slot_id == slot_id
        ).order_by(DBSlotValue.sort_order.asc()).all()
        
        return success_response(data={"slot_values": slot_values})
    except HTTPException:
        raise
    except Exception as e:
        return error_response(msg=f"获取词槽值失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR)


@router.post("/{slot_id}/values", response_model=StandardResponse)
async def add_slot_value(
    slot_id: int,
    value_data: SlotValueBase,
    db: Session = Depends(get_new_db)
):
    """添加词槽值"""
    try:
        # 验证词槽是否存在
        slot = db.query(DBSlotDefinition).filter(
            DBSlotDefinition.id == slot_id
        ).first()
        
        if not slot:
            return error_response(msg="词槽不存在", code=ErrorCodes.NOT_FOUND)
        
        # 检查标准值是否已存在
        existing_value = db.query(DBSlotValue).filter(
            DBSlotValue.slot_id == slot_id,
            DBSlotValue.standard_value == value_data.standard_value
        ).first()
        
        if existing_value:
            return error_response(msg="标准值已存在", code=ErrorCodes.PARAM_ERROR)
        
        # 获取最大排序序号
        max_order = db.query(func.max(DBSlotValue.sort_order)).filter(
            DBSlotValue.slot_id == slot_id
        ).scalar() or 0
        
        # 创建词槽值 - 默认启用状态
        new_value = DBSlotValue(
            slot_id=slot_id,
            standard_value=value_data.standard_value,
            aliases=value_data.aliases,
            description=value_data.description,
            is_active=True,  # 固定为启用状态
            sort_order=max_order + 1
        )
        
        db.add(new_value)
        db.commit()
        db.refresh(new_value)
        
        return success_response(data={"slot_value": new_value}, msg="词槽值添加成功")
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return error_response(msg=f"添加词槽值失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR)


@router.put("/{slot_id}/values/{value_id}", response_model=StandardResponse)
async def update_slot_value(
    slot_id: int,
    value_id: int,
    value_data: SlotValueBase,
    db: Session = Depends(get_new_db)
):
    """更新词槽值"""
    try:
        # 验证词槽是否存在
        slot = db.query(DBSlotDefinition).filter(
            DBSlotDefinition.id == slot_id
        ).first()
        
        if not slot:
            return error_response(msg="词槽不存在", code=ErrorCodes.NOT_FOUND)
        
        # 获取词槽值
        slot_value = db.query(DBSlotValue).filter(
            DBSlotValue.id == value_id,
            DBSlotValue.slot_id == slot_id
        ).first()
        
        if not slot_value:
            return error_response(msg="词槽值不存在", code=ErrorCodes.NOT_FOUND)
        
        # 检查标准值是否与其他词槽值冲突
        existing_value = db.query(DBSlotValue).filter(
            DBSlotValue.slot_id == slot_id,
            DBSlotValue.standard_value == value_data.standard_value,
            DBSlotValue.id != value_id
        ).first()
        
        if existing_value:
            return error_response(msg="标准值已存在", code=ErrorCodes.PARAM_ERROR)
        
        # 更新词槽值 - 保持启用状态
        slot_value.standard_value = value_data.standard_value
        slot_value.aliases = value_data.aliases
        slot_value.description = value_data.description
        # slot_value.is_active 保持不变，不更新状态字段
        
        db.commit()
        db.refresh(slot_value)
        
        return success_response(data={"slot_value": slot_value}, msg="词槽值更新成功")
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return error_response(msg=f"更新词槽值失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR)


@router.delete("/{slot_id}/values/{value_id}", response_model=StandardResponse)
async def delete_slot_value(
    slot_id: int,
    value_id: int,
    db: Session = Depends(get_new_db)
):
    """删除词槽值"""
    try:
        # 验证词槽是否存在
        slot = db.query(DBSlotDefinition).filter(
            DBSlotDefinition.id == slot_id
        ).first()
        
        if not slot:
            return error_response(msg="词槽不存在", code=ErrorCodes.NOT_FOUND)
        
        # 获取词槽值
        slot_value = db.query(DBSlotValue).filter(
            DBSlotValue.id == value_id,
            DBSlotValue.slot_id == slot_id
        ).first()
        
        if not slot_value:
            return error_response(msg="词槽值不存在", code=ErrorCodes.NOT_FOUND)
        
        # 删除词槽值
        db.delete(slot_value)
        db.commit()
        
        return success_response(msg="词槽值删除成功")
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        return error_response(msg=f"删除词槽值失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR)


@router.get("/{slot_id}/related-instructions", response_model=StandardResponse)
async def get_related_instructions(
    slot_id: int,
    db: Session = Depends(get_new_db)
):
    """获取词槽关联的指令"""
    try:
        # 验证词槽是否存在
        slot = db.query(DBSlotDefinition).filter(
            DBSlotDefinition.id == slot_id
        ).first()
        
        if not slot:
            return error_response(msg="词槽不存在", code=ErrorCodes.NOT_FOUND)
        
        # 获取关联指令
        related_instructions = db.query(DBInstructionData).filter(
            DBInstructionData.related_slot_ids.contains(f'[{slot_id}]') |
            DBInstructionData.related_slot_ids.contains(f'{slot_id}')
        ).all()
        
        return success_response(data={"instructions": related_instructions})
    except HTTPException:
        raise
    except Exception as e:
        return error_response(msg=f"获取关联指令失败: {str(e)}", code=ErrorCodes.SYSTEM_ERROR) 